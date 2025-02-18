import os
import gc
import csv
import warnings
import pandas as pd

from multiprocessing import Pool, freeze_support

from openpyxl import Workbook, load_workbook


class Book:
    def _cuts_numbers(self, text: str) -> str:
        """Обрезает цифры, оставляет буквы для Ячеек"""
        return "".join([char for char in text if not char.isdigit()])

    def reed_book(self, filename: str) -> list[dict[str]]:
        # Убираем предупреждение
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")

            wb = load_workbook(filename=filename, read_only=True)
            # read_only — режим для отложенной загрузки, экономить оперативную память

            sheet_ranges = wb.active

        book = {}
        organization = sheet_ranges["B25"].value

        order = []
        for coll in sheet_ranges["C32":"V32"]:
            for cell in coll:
                order.append(str(cell.value))
        order = "".join(order)

        id_cell = 43  # Счетчик для row обработки таблицы
        book_csv = []
        while True:
            id_cell += 1
            id_coll_start = f"A{id_cell}"

            if sheet_ranges[id_coll_start].value is not None:
                book["№"] = sheet_ranges[f"A{id_cell}"].value
                book["дата_старт"] = sheet_ranges[f"B{id_cell}"].value
                book["дата_end"] = sheet_ranges[f"E{id_cell}"].value
                book["вид"] = sheet_ranges[f"H{id_cell}"].value
                book["номер"] = sheet_ranges[f"J{id_cell}"].value
                book["дата_реквизита"] = sheet_ranges[f"M{id_cell}"].value
                book["номер_кор"] = sheet_ranges[f"P{id_cell}"].value
                book["наименование"] = sheet_ranges[f"S{id_cell}"].value
                book["бик"] = sheet_ranges[f"V{id_cell}"].value
                book["фио"] = sheet_ranges[f"Y{id_cell}"].value
                book["инн"] = sheet_ranges[f"AB{id_cell}"].value
                book["кпп"] = sheet_ranges[f"AE{id_cell}"].value
                book["номер_счета"] = sheet_ranges[f"AH{id_cell}"].value
                book["дебет"] = sheet_ranges[f"AK{id_cell}"].value
                book["кредит"] = sheet_ranges[f"AN{id_cell}"].value
                book["назначение"] = sheet_ranges[f"AQ{id_cell}"].value
                book["ордер"] = order
                book["организация"] = organization
                book_csv.append(book)
                book = {}
            else:
                gc.collect()  # Чистим мусор
                wb.close()  # Закрываем книгу
                break

        return book_csv

    def create_new_book(self, data: dict[str], file_name: str) -> None:
        """Создает книгу для xlsx"""
        wb_new = Workbook()

        sheet = wb_new.active
        sheet.title = "Сводный отчет"
        for k, v in data.items():
            sheet[k] = v
        wb_new.save(file_name)
        wb_new.close()

    def save_book_csv(self, data: list[dict[str]], file_name: str) -> None:
        with open(f"{file_name}.csv", "w", newline="", encoding="utf-8") as csv_file:
            writer = csv.DictWriter(csv_file, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)


def _find_files(directory: str, extension: str) -> list:
    """Формирует список path где лежат файлы excel"""
    return [
        os.path.join(directory, file)
        for file in os.listdir(directory)
        if file.endswith(extension)
    ]


def create_directory(name: str) -> None:
    os.makedirs(name, exist_ok=True)


def read_csv_files(directory: str) -> pd.DataFrame:
    """Читает все CSV файлы в директории и возвращает объединенный DataFrame."""
    combined_data = pd.DataFrame()  # Пустой DataFrame для объединенных данных
    for file_path in _find_files(directory, extension=".csv"):
        df = pd.read_csv(file_path, encoding="utf-8")  # Чтение CSV файла
        combined_data = pd.concat(
            [combined_data, df], ignore_index=True
        )  # Объединение данных
    return combined_data


def save_to_xlsx(data: pd.DataFrame, output_filename: str) -> None:
    """Сохраняет DataFrame в XLSX файл."""
    with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
        data.to_excel(writer, index=False, sheet_name="Combined Data")
    print(f"Данные сохранены в файл: {output_filename}")


def merge_csv_to_xlsx(input_directory: str, output_filename: str) -> None:
    """Объединяет все CSV файлы в директории в один XLSX файл."""
    combined_data = read_csv_files(input_directory)
    save_to_xlsx(combined_data, output_filename)  # Сохранение в XLSX


def delete_csv_files_directory(name_directory: str):
    os.unlink()


def worker(path: str, name_directory: str) -> str | None:
    try:
        bk = Book()
        data = bk.reed_book(path)
        path_directory = os.path.join(name_directory, os.path.basename(path))
        # bk.create_new_book(data, file_name=path_directory)
        bk.save_book_csv(data, file_name=path_directory)
        return f"Файл создан: {path_directory}"
    except Exception as e:
        return f"Ошибка при обработке файла {path}: {e}"


def clear_csv_files(directory: str) -> None:
    """
    Удаляет все CSV файлы в указанной директории.

    :param directory: Путь к директории, в которой нужно очистить CSV файлы.
    """
    if not os.path.exists(directory):
        print(f"Директория {directory} не существует.")
        return

    # Получаем список всех файлов в директории
    files = os.listdir(directory)

    # Фильтруем только CSV файлы
    csv_files = [file for file in files if file.endswith(".csv")]

    if not csv_files:
        print(f"В директории {directory} нет CSV файлов для удаления.")
        return

    # Удаляем каждый CSV файл
    for csv_file in csv_files:
        file_path = os.path.join(directory, csv_file)
        try:
            os.remove(file_path)
            print(f"Файл {csv_file} успешно удален.")
        except Exception as e:
            print(f"Ошибка при удалении файла {csv_file}: {e}")


def main() -> None:
    name_directory = "00_Data"
    create_directory(name_directory)
    # file_list = find_files(os.getcwd(), extension="csv")
    file_list = _find_files(r"C:\Projects\auto_exc\example_files", extension=".xlsx")

    print(f"Всего файлов в директории {len(file_list)}. Идет обработка файлов:")
    total_files = len(file_list)
    processed_files = 1
    with Pool(os.cpu_count()) as pool:
        results = []
        results = pool.starmap(worker, [(path, name_directory) for path in file_list])

        for result in results:
            print(f"Обработано файлов: {processed_files}/{total_files}")
            processed_files += 1
            print(result)
    print("Обработка файлов завершена успешна. Следующий этап — Объединение данных")
    # os.system('cls||clear')

    output_filename = os.path.join(name_directory, "combined_data.xlsx")
    merge_csv_to_xlsx(name_directory, output_filename)

    print("Следующий этап — Очистка файлов")
    clear_csv_files(name_directory)
    # Пауза, чтобы консоль не закрывалась
    input("\nНажмите Enter для выхода...")


if __name__ == "__main__":
    freeze_support()
    main()
