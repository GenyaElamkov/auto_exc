import os
import gc
import threading


from tqdm import tqdm
from openpyxl import Workbook, load_workbook


def find_xlsx_files(directory: str) -> list:
    """Формирует список path где лежат файлы excel"""
    return [
        f"{directory}\\{file}"
        for file in os.listdir(directory)
        if file.endswith(".xlsx")
    ]


def create_directory(name: str):
    os.makedirs(name, exist_ok=True)


class Book:
    def __init__(self) -> None:
        self.counter_row = 1  # Счетчик для row

    def _cuts_numbers(self, text: str):
        """Обрезает цифры, оставляет буквы для Ячеек"""
        return "".join([char for char in text if not char.isdigit()])

    def reed_book(self, filename: str) -> dict[str]:
        wb = load_workbook(filename=filename, read_only=True)
        # read_only — режим для отложенной загрузки, экономить оперативную память
        sheet_ranges = wb.active
        
        book = {}
        organization = sheet_ranges['B25'].value
        
        order = []
        for coll in sheet_ranges["C32":"V32"]:
            for cell in coll:
                order.append(cell.value)
        order = "".join(order)

        id_cell = 43        # Счетчик для row обработки таблицы
        counter_col = 0     # Счетчик для col
        while True:
            id_cell += 1
            id_coll_start = f"A{id_cell}"
            id_coll_end = f"AV{id_cell}"

            if sheet_ranges[id_coll_start].value is not None:
                for coll in sheet_ranges[id_coll_start:id_coll_end]:
                    for cell in coll:
                        coordinate = self._cuts_numbers(cell.coordinate)
                        book.setdefault(f"{coordinate}{self.counter_row}", cell.value)
                        counter_col += 1
                    book.setdefault(f"AW{self.counter_row}", order)
                    book.setdefault(f"AX{self.counter_row}", organization)
                    counter_col = 0
                self.counter_row += 1
            else:
                self.counter_row = 1
                gc.collect()  # Чистим мусор
                wb.close()  # Закрываем книгу
                break
        return book

    def create_new_book(self, data: dict[str], file_name: str) -> None:
        wb_new = Workbook()

        sheet = wb_new.active
        sheet.title = "Сводный отчет"
        for k, v in data.items():
            sheet[k] = v
        wb_new.save(file_name)
        wb_new.close()


def main(path, name_directory):
    bk = Book()
    data = bk.reed_book(path)
    path_directory = f"{name_directory}\\{path.split("\\")[-1]}"
    bk.create_new_book(data, file_name=path_directory)


if __name__ == "__main__":
    try:
        name_directory = "00_Data"
        create_directory(name_directory)
        file_list = find_xlsx_files(os.getcwd())
        theards = []
        for path in tqdm(file_list):
            th = threading.Thread(target=main, args=(path, name_directory), daemon=True)
            th.start()
            theards.append(th)

        for th in tqdm(theards):
            th.join()

    except Exception as e:
        print(f"Ошибка {e}")
