from calendar import c
import os

import re
from string import ascii_uppercase
from tkinter import NO
from openpyxl import Workbook, load_workbook
from requests_toolbelt import NonMultipartContentTypeException


def get_value_cells():
    pass


def find_xlsx_files(directory: str) -> list:
    xlsx_files = []

    for file in os.listdir(directory):
        if file.endswith(".xlsx"):
            xlsx_files.append(f"{directory}\{file}")
    return xlsx_files







class Book:
    def __init__(self) -> None:
        self.counter_row = 1

    def cuts_numbers(self, text:str):
        chars = ''
        for i in text:
            if not i.isdigit():
                chars += i
        return chars


    def set_book(self, filename: str) -> dict[str]:
        wb = load_workbook(filename=filename)
        sheet_ranges = wb.active
        order = []
        for coll in sheet_ranges["C32":"V32"]:
            for cell in coll:
                if not cell.value:
                    continue
                order.append(cell.value)
        order = "".join(order)

        id_cell = 43
        counter_col = 0  # Счетчик для col
        # self.counter_row = 1 # Счетчик для row
        book = {}

        while True:
            id_cell += 1
            id_coll_start = f"A{id_cell}"
            id_coll_end = f"AV{id_cell}"

            if sheet_ranges[id_coll_start].value is not None:
                for coll in sheet_ranges[id_coll_start:id_coll_end]:
                    for cell in coll:
                        # if not cell.value:
                        #     continue
                        coordinate = self.cuts_numbers(cell.coordinate)
                        book.setdefault(f"{coordinate}{self.counter_row}", cell.value)
                        # book.setdefault(
                        #     f"{ascii_uppercase[counter_col]}{self.counter_row}", cell.value
                        # )
                        counter_col += 1
                    # book.setdefault(
                    #     f"{ascii_uppercase[counter_col]}{self.counter_row}", order
                    # )
                    book.setdefault(f"AW{self.counter_row}", order)
                        
                    counter_col = 0
                self.counter_row += 1
            else:
                break
        return book
    

    def create_new_book(self, data: dict[str]) -> None:
        wb_new = Workbook()
        sheet = wb_new.active
        sheet.title = "Сводный отчет"
        for k, v in data.items():
            sheet[k] = v

        wb_new.save("report.xlsx")




def main():
    data = {}

    bk = Book()
    for path in find_xlsx_files(r"C:\Projects\auto_exc\data"):
        data.update(bk.set_book(path))
    bk.create_new_book(data)


if __name__ == "__main__":
    main()
