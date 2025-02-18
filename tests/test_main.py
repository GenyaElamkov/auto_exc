import os
import sys
import tempfile
from openpyxl import Workbook, load_workbook

# Добавляем корневую директорию в sys.path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from project.auto_exc import main


def test_main(capsys):
    # Создаем временную директорию с тестовыми файлами
    with tempfile.TemporaryDirectory() as temp_dir:
        # Создаем несколько Excel-файлов
        for i in range(3):
            wb = Workbook()
            ws = wb.active
            ws["B25"] = "Org"
            wb.save(os.path.join(temp_dir, f"test{i}.xlsx"))

        # Меняем текущую рабочую директорию на временную
        original_dir = os.getcwd()
        os.chdir(temp_dir)

        try:
            # Запускаем main
            main()
            captured = capsys.readouterr()
            assert "Обработано файлов" in captured.out
            # Проверяем, что папка 00_Data создана
            output_dir = os.path.join(temp_dir, "00_Data")
            assert os.path.exists(output_dir)

            # Проверяем, что файлы созданы
            output_files = os.listdir(output_dir)
            assert len(output_files) == 3

        finally:
            # Возвращаемся в исходную директорию
            os.chdir(original_dir)
