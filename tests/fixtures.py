import csv
import pytest

from openpyxl import Workbook


# Фикстуры для тестовых данных
@pytest.fixture
def tmp_excel_file(tmp_path):
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    # Заполняем тестовые данные согласно структуре скрипта
    ws["B25"] = "Test Organization"
    ws["C32"] = "1"
    ws["D32"] = "1"
    ws["E32"] = "1"
    ws["F32"] = "1"
    ws["G32"] = "1"
    ws["H32"] = "1"
    ws["I32"] = "1"
    ws["J32"] = "1"
    ws["K32"] = "1"
    ws["L32"] = "1"
    ws["M32"] = "1"
    ws["N32"] = "1"
    ws["O32"] = "1"
    ws["P32"] = "1"
    ws["Q32"] = "1"
    ws["R32"] = "1"
    ws["S32"] = "1"
    ws["T32"] = "1"
    ws["U32"] = "1"
    ws["V32"] = "1"
    # Заполняем строку данных (ячейка A43+)
    ws["A44"] = 1
    ws["B44"] = "2023-01-01"
    ws["E44"] = "2023-01-02"
    ws["H44"] = "TypeA"
    ws["M44"] = "123"
    ws["P44"] = "456"
    ws["S44"] = "Test Name"
    ws["V44"] = "123456789"
    ws["Y44"] = "Иванов И.И."
    ws["AB44"] = "1234567890"
    ws["AE44"] = "987654321"
    ws["AH44"] = "40702810000000012345"
    ws["AK44"] = "100.50"
    ws["AN44"] = "200.75"
    ws["AQ44"] = "Назначение платежа"
    wb.save(file_path)
    return file_path


@pytest.fixture
def tmp_csv_files(tmp_path):
    csv_dir = tmp_path / "csv_dir"
    csv_dir.mkdir()
    # Создаем 2 тестовых CSV файла
    data1 = [{"№": 1, "дата_старт": "2023-01-01"}, {"№": 2, "дата_старт": "2023-01-02"}]
    data2 = [{"№": 3, "дата_старт": "2023-01-03"}]

    with open(csv_dir / "file1.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=data1[0].keys())
        writer.writeheader()
        writer.writerows(data1)

    with open(csv_dir / "file2.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=data2[0].keys())
        writer.writeheader()
        writer.writerows(data2)

    return csv_dir
