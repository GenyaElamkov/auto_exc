import os
import sys
import tempfile
from openpyxl import Workbook

# Добавляем корневую директорию в sys.path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from project.auto_exc import Book


def test_reed_book():
    # Создаем временный Excel-файл
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
        wb = Workbook()
        ws = wb.active

        # Заполняем тестовыми данными
        ws["B25"] = "Test Organization"
        for key in [
            "C",
            "D",
            "E",
            "F",
            "G",
            "H",
            "I",
            "J",
            "K",
            "L",
            "M",
            "N",
            "O",
            "P",
            "Q",
            "R",
            "S",
            "T",
            "U",
            "V",
        ]:
            ws[f"{key}32"] = "1"

        ws["A44"] = "Data1"
        # ws["B44"] = "Data2"
        # ws["H44"] = "Data3"
        # ws["J44"] = "Data4"
        # ws["M44"] = "Data5"
        # ws["V44"] = "Data6"
        # ws["Y44"] = "Data7"
        # ws["AB44"] = "Data8"
        # ws["AE44"] = "Data9"
        # ws["AH44"] = "Data10"
        # ws["AK44"] = "Data11"
        # ws["AN44"] = "Data12"
        # ws["AQ44"] = "Data13"

        wb.save(tmp_file.name)

    # Тестируем метод reed_book
    book = Book()
    result = book.reed_book(tmp_file.name)

    # Проверяем результат
    assert result["R1"] == "Test Organization"
    assert result["A1"] == "Data1"
    # assert result["B1"] == "Data2"
    # assert result["H1"] == "Data3"
    # assert result["J1"] == "Data4"
    # assert result["M1"] == "Data5"
    # assert result["V1"] == "Data6"
    # assert result["Y1"] == "Data7"
    # assert result["AB1"] == "Data8"
    # assert result["AE1"] == "Data9"
    # assert result["AH1"] == "Data10"
    # assert result["AK1"] == "Data11"
    # assert result["AN1"] == "Data12"
    # assert result["AQ1"] == "Data13"
    assert result["Q1"] == "11111111111111111111"
    # assert "C1" not in result

    # Удаляем временный файл
    os.unlink(tmp_file.name)
