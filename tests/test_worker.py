import os
import sys
import tempfile

from openpyxl import Workbook

# Добавляем корневую директорию в sys.path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from project.auto_exc import worker, Book


def test_worker():
    # Создаем временный Excel-файл
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
        wb = Workbook()
        ws = wb.active

        # Заполняем тестовыми данными
        ws["B25"] = "Test Organization"
        ws["C32"] = "Order1"
        ws["A43"] = "Data1"
        ws["B43"] = "Data2"
        wb.save(tmp_file.name)

    # Создаем временную директорию для результатов
    with tempfile.TemporaryDirectory() as temp_dir:
        print(f"Временная директория: {temp_dir}")  # Отладочный вывод

        # Вызываем функцию worker
        result = worker(tmp_file.name, temp_dir)
        # Проверяем, что файл создан
        output_file = os.path.join(temp_dir, os.path.basename(tmp_file.name))
        print(f"Ожидаемый путь к файлу: {output_file}")  # Отладочный вывод
        assert os.path.exists(output_file), f"Файл {output_file} не найден"

        # Проверяем сообщение об успехе
        assert (
            "Файл создан" in result
        ), f"Ожидалось сообщение об успехе, но получено: {result}"

    # Удаляем временный файл
    os.unlink(tmp_file.name)
